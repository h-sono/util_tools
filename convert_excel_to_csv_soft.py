# 配布用(exeファイル)
import tkinter
from tkinter import *
# ttkはGUI定義に使用
from tkinter import ttk
from tkinter import filedialog
import os
import csv
import openpyxl
import pandas as pd
import re

def click_refer_button():
  """ ファイルの参照処理 """
  # ファイルタイプは特に指定なし。指定する場合、("HTMLファイル", "*.html")のようにリスト形式で指定。
  filetypes = [("","*")]
  initialdir = os.path.abspath(os.path.dirname(__file__))
  get_file_path = filedialog.askopenfilename(filetypes = filetypes, initialdir = initialdir)
  # 取得したファイルパスを変数file_pathにセットし、参照したファイルパス表示に使用。
  file_path.set(get_file_path)

  return get_file_path

def get_excel_sheet() -> list:
  """ エクセルシート取得 """  
  get_file_path = click_refer_button
  # ファイル名を取得
  match = re.search(r"[^\\ | \/]+$", get_file_path)
  # エクセルファイル読み込み
  work_book = openpyxl.load_workbook(match.group())
  # シート名(リスト)を返却
  return work_book.sheetnames

# CSVファイル出力処理
def click_export_button(file_name: str, sheet_num: int):
  # シート名をリストで取得
  input_book = pd.ExcelFile(file_name)
  input_sheet_name = input_book.sheet_names

  # エクセルファイル読み込み
  work_book = openpyxl.load_workbook(file_name)
  # 特定シートを指定(シートの番号をインデックス番号で指定)
  work_sheet = work_book[input_sheet_name[sheet_num]]

  # シートの行データを取得(iter_rowsで行を読み取る)
  all_rows_list = []
  for row in work_sheet.iter_rows():
    row_list = []
    for cell in row:
        # IntegerFieldのカラムにはエクセルファイル側で事前に数値を入れる必要あり。
        # それ以外の空白セルは文字列型のカラムと判断し、Noneが来たら空文字とする。
        if cell.value is None:
            cell.value = ""
            row_list.append(str(cell.value))
        else:
            row_list.append(str(cell.value))
    all_rows_list.append(row_list)

  # csvファイルへの書き込み
  with open("convert_excel_result.csv", "w", encoding="UTF-8") as f:
      writer = csv.writer(f)
      for item in all_rows_list:
          writer.writerow(item)


if __name__ == '__main__':
  # Tkinterインスタンスの生成
  root = tkinter.Tk()
  # ウィンドウタイトル
  root.title("エクセル/CSV変換ツール")
  # ウィンドウ初期表示サイズ
  root.geometry("800x600")

  # Frame1の作成(ttk.Frameは複数のウィジェットを取りまとめる)
  # 「ファイル名：」、参照ファイルパスを表示するテキストボックス、参照ボタンをまとめるFrame
  frame_1 = ttk.Frame(root, padding=10)
  frame_1.grid()

  # 「ファイル名：」ラベルの作成。
  s = StringVar()
  s.set('ファイル名：')
  label_1 = ttk.Label(frame_1, textvariable=s)
  label_1.grid(row=0, column=0)

  # 参照ファイルのパスを表示するテキストボックスの作成(ttk.Entry)。
  file_path = StringVar()
  filepath_entry = ttk.Entry(frame_1, textvariable=file_path, width=50)
  # row=0とし、「ファイル名：」と同一行で2列目に配置する。
  filepath_entry.grid(row=0, column=1)

  # 参照ボタンの作成。
  refer_button = ttk.Button(frame_1, text=u'参照', command=click_refer_button)
  # row=0とし、「ファイル名：」と同一行で3列目に配置する。
  refer_button.grid(row=0, column=2)
  refer_button.bind("<ButtonPress>", get_excel_sheet())

  # Frame2の作成
  frame_2 = ttk.Frame(root, padding=10)
  frame_2.grid()

  # エクセルシート選択
  combobox = ttk.Combobox(root, height=3, values=get_excel_sheet())
  combobox.grid(row=1, column=0)

  # ウィンドウの描画
  root.mainloop()
